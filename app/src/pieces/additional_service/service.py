import io
import os
from tempfile import SpooledTemporaryFile
from typing import Union

from openpyxl.reader.excel import load_workbook
from sqlalchemy import func, text

from app.src.config import DATA_FOLDER_PATH
from app.src.pieces.additional_service.models import AdditionalServiceModel
from app.src.pieces.additional_service.schemas import AdditionalServiceCreationSchema
from sqlalchemy.orm import Session


def refresh_table(db: Session):
    db.execute(text("TRUNCATE TABLE additional_service"))


def get_additional_service_by_id(db: Session, additional_service_id: int) -> AdditionalServiceModel:
    return db.query(AdditionalServiceModel) \
        .filter(AdditionalServiceModel.id == additional_service_id).first()


def get_additional_services(db: Session, skip: int = 0, limit: int = 100) -> list[AdditionalServiceModel]:
    return db.query(AdditionalServiceModel).offset(skip).limit(limit).all()


def get_additional_service_suggestions(db: Session, subtext: str = '',
                                       skip: int = 0, limit: int = 100) -> list[AdditionalServiceModel]:
    if subtext == '':
        return get_additional_services(db, skip, limit)
    return db.query(AdditionalServiceModel) \
        .filter(func.lower(AdditionalServiceModel.name).contains(subtext.lower())).offset(skip).limit(limit).all()


def add_additional_service(db: Session,
                           additional_service: AdditionalServiceCreationSchema) -> AdditionalServiceModel:
    additional_service = AdditionalServiceModel(**additional_service.dict())
    db.add(additional_service)
    db.commit()
    db.refresh(additional_service)
    return additional_service


def update_additional_service(db: Session, id: int,
                              schema: AdditionalServiceCreationSchema) -> AdditionalServiceModel:
    additional_service = db.query(AdditionalServiceModel) \
        .filter(AdditionalServiceModel.id == id).first()
    additional_service.average_price_dollar = schema.average_price_dollar
    additional_service.name = schema.name
    db.commit()
    return additional_service


def delete_additional_service(db: Session, id: int,) -> AdditionalServiceModel:
    additional_service = db.query(AdditionalServiceModel) \
        .filter(AdditionalServiceModel.id == id).first()
    db.delete(additional_service)
    db.commit()
    return additional_service


async def upload_additional_services_excel_to_db(file: SpooledTemporaryFile, refresh: bool, db: Session):
    if refresh:
        refresh_table(db)

    f = await file.read()
    xlsx = io.BytesIO(f)
    workbook = load_workbook(xlsx, data_only=True)
    worksheet = workbook.worksheets[0]

    for row in worksheet.iter_rows(min_row=2):
        schema = AdditionalServiceCreationSchema(name=row[0].value, average_price_dollar=float(row[1].value))
        add_additional_service(db, schema)


def parse_additional_services(filename: str, db: Session, only_first: Union[int, None] = None):
    if only_first is not None:
        only_first += 2

    file_path = os.path.join(DATA_FOLDER_PATH, filename)
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, max_row=only_first, max_col=5, values_only=True):
        if row[2] is None or row[2] == '-':
            continue
        schema = AdditionalServiceCreationSchema(name=row[1], average_price_dollar=float(row[2]))
        res = add_additional_service(db, schema)
