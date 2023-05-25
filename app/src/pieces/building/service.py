import io
from tempfile import SpooledTemporaryFile

from sqlalchemy import func, text
from sqlalchemy.orm import Session

from app.src.config import DATA_FOLDER_PATH
from app.src.pieces.building.models import BuildingModel
from app.src.pieces.building.schemas import BuildingCreationSchema
import os
from typing import Union

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.src.pieces.building.schemas import BuildingCreationSchema


def refresh_table(db: Session):
    db.execute(text("TRUNCATE TABLE building"))


def get_building_by_id(db: Session, user_id: int) -> BuildingModel:
    return db.query(BuildingModel).filter(BuildingModel.id == user_id).first()


def get_buildings(db: Session, skip: int = 0, limit: int = 100) -> list[BuildingModel]:
    return db.query(BuildingModel).offset(skip).limit(limit).all()


def get_building_suggestions(db: Session, subtext: str = '', skip: int = 0, limit: int = 100) -> list[BuildingModel]:
    if subtext == '':
        return get_buildings(db, skip, limit)
    return db.query(BuildingModel)\
        .filter(func.lower(BuildingModel.name).contains(subtext.lower())).offset(skip).limit(limit).all()


def add_building(db: Session, building: BuildingCreationSchema) -> BuildingModel:
    building = BuildingModel(**building.dict())
    db.add(building)
    db.commit()
    db.refresh(building)
    return building


def update_building(db: Session, id: int,
                    schema: BuildingCreationSchema) -> BuildingModel:
    building = db.query(BuildingModel) \
        .filter(BuildingModel.id == id).first()
    building.average_price_rub = schema.average_price_rub
    building.name = schema.name
    db.commit()
    return building


def delete_building(db: Session, id: int, ) -> BuildingModel:
    building = db.query(BuildingModel).filter(BuildingModel.id == id).first()
    db.delete(building)
    db.commit()
    return building


async def upload_building_excel_to_db(file: SpooledTemporaryFile, refresh: bool, db: Session):
    if refresh:
        refresh_table(db)

    f = await file.read()
    xlsx = io.BytesIO(f)
    workbook = load_workbook(xlsx, data_only=True)
    worksheet = workbook.worksheets[0]

    for row in worksheet.iter_rows(min_row=2):
        schema = BuildingCreationSchema(name=row[0].value, average_price_rub=float(row[1].value))
        add_building(db, schema)


def parse_buildings(filename: str, db: Session, only_first: Union[int, None] = None):
    if only_first is not None:
        only_first += 2

    file_path = os.path.join(DATA_FOLDER_PATH, filename)
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, max_row=only_first, max_col=5, values_only=True):
        if row[2] is None or row[2] == '-':
            continue
        schema = BuildingCreationSchema(name=row[1], average_price_rub=float(row[2]))
        res = add_building(db, schema)
        # print('eq created')
        # print(res.average_price_dollar)
        # print(res.name)
        # print(res.id)
