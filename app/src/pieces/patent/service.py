from app.src.config import DATA_FOLDER_PATH
import os
from typing import Union

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.src.pieces.patent.models import PatentModel
from app.src.pieces.patent.schemas import PatentCreationSchema


def create_patent(patent: PatentCreationSchema, db: Session) -> PatentModel:
    patent = PatentModel(**patent.dict())
    db.add(patent)
    db.commit()
    db.refresh(patent)
    return patent


def parse_patents(filename: str, db: Session, only_first: Union[int, None] = None):
    if only_first is not None:
        only_first += 3
    file_path = os.path.join(DATA_FOLDER_PATH, filename)

    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=3, max_row=only_first, max_col=5, values_only=True):
        #print(row)

        if None in (row[1], row[2], row[3]):
            continue

        if db.query(PatentModel).filter(PatentModel.name == row[1]).first() is not None:
            continue

        schema = PatentCreationSchema(name=row[1], income_rub=float(row[2]), percent_rate=float(row[3]), price=float(row[2]) * float(row[3]) / 100)
        res = create_patent(schema, db)
        #print('eq created')
        #print(schema.name, schema.income_rub, schema.percent_rate, schema.price)
        #print('-----')
        #print(res.average_price_dollar)
        #print(res.name)
        #print(res.id)  # todo check it

