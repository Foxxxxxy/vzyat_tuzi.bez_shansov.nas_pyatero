import io
from tempfile import SpooledTemporaryFile

from sqlalchemy import func, text

from app.src.config import DATA_FOLDER_PATH
from app.src.pieces.equipment.models import EquipmentModel
import os
from typing import Union

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.src.pieces.equipment.schemas import EquipmentCreationSchema


def refresh_table(db: Session):
    db.execute(text("TRUNCATE TABLE equipment"))


def get_equipment_by_id(db: Session, equipment_id: int) -> EquipmentModel:
    return db.query(EquipmentModel).filter(EquipmentModel.id == equipment_id).first()


def get_equipments(db: Session, skip: int = 0, limit: int = 100) -> list[EquipmentModel]:
    return db.query(EquipmentModel).offset(skip).limit(limit).all()


def get_equipment_suggestions(db: Session, subtext: str = '', skip: int = 0, limit: int = 100) -> list[EquipmentModel]:
    if subtext == '':
        return get_equipments(db, skip, limit)
    return db.query(EquipmentModel)\
        .filter(func.lower(EquipmentModel.name).contains(subtext.lower())).offset(skip).limit(limit).all()


def add_equipment(db: Session, equipment: EquipmentCreationSchema) -> EquipmentModel:
    equipment = EquipmentModel(**equipment.dict())
    db.add(equipment)
    db.commit()
    db.refresh(equipment)
    return equipment


def update_equipment(db: Session, id: int,
                     schema: EquipmentCreationSchema) -> EquipmentModel:
    equipment = db.query(EquipmentModel) \
        .filter(EquipmentModel.id == id).first()
    equipment.average_price_dollar = schema.average_price_dollar
    equipment.name = schema.name
    db.commit()
    return equipment


def delete_equipment(db: Session, id: int, ) -> EquipmentModel:
    equipment = db.query(EquipmentModel).filter(EquipmentModel.id == id).first()
    db.delete(equipment)
    db.commit()
    return equipment


async def upload_equipment_excel_to_db(file: SpooledTemporaryFile, refresh: bool, db: Session):
    if refresh:
        refresh_table(db)

    f = await file.read()
    xlsx = io.BytesIO(f)
    workbook = load_workbook(xlsx, data_only=True)
    worksheet = workbook.worksheets[0]

    for row in worksheet.iter_rows(min_row=2):
        schema = EquipmentCreationSchema(name=row[0].value, average_price_dollar=float(row[1].value))
        add_equipment(db, schema)


def parse_stanki(filename: str, db: Session, only_first: Union[int, None] = None):
    if only_first is not None:
        only_first += 2

    file_path = os.path.join(DATA_FOLDER_PATH, filename)
    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, max_row=only_first, max_col=5, values_only=True):
        if row[2] is None or row[2] == '-':
            continue
        schema = EquipmentCreationSchema(name=row[1], average_price_dollar=float(row[2]))
        res = add_equipment(db, schema)
        # print('eq created')
        # print(res.average_price_dollar)
        # print(res.name)
        # print(res.id)
