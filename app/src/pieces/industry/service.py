import io
from tempfile import SpooledTemporaryFile

from sqlalchemy import func, text

from app.src.config import DATA_FOLDER_PATH
from app.src.pieces.industry.models import IndustryModel
from app.src.pieces.industry.schemas import IndustryCreationSchema
import os
from typing import Union

from openpyxl import load_workbook
from sqlalchemy.orm import Session


def refresh_table(db: Session):
    db.execute(text("TRUNCATE TABLE industry"))


def get_industry_by_id(db: Session, user_id: int) -> IndustryModel:
    return db.query(IndustryModel).filter(IndustryModel.id == user_id).first()


def get_industries(db: Session, skip: int = 0, limit: int = 100) -> list[IndustryModel]:
    return db.query(IndustryModel).offset(skip).limit(limit).all()


def get_industry_suggestions(db: Session, subtext: str = '', skip: int = 0, limit: int = 100) -> list[IndustryModel]:
    if subtext == '':
        return get_industries(db, skip, limit)
    return db.query(IndustryModel)\
        .filter(func.lower(IndustryModel.name).contains(subtext.lower())).offset(skip).limit(limit).all()


def add_industry(db: Session, industry: IndustryCreationSchema) -> IndustryModel:
    industry = IndustryModel(**industry.dict())
    db.add(industry)
    db.commit()
    db.refresh(industry)
    return industry


def update_industry(db: Session, id: int,
                    schema: IndustryCreationSchema) -> IndustryModel:
    industry = db.query(IndustryModel) \
        .filter(IndustryModel.id == id).first()
    industry.name = schema.name
    industry.avg_salary = schema.avg_salary
    db.commit()
    return industry


def delete_industry(db: Session, id: int) -> IndustryModel:
    industry = db.query(IndustryModel) \
        .filter(IndustryModel.id == id).first()
    db.delete(industry)
    db.commit()
    return industry


async def upload_industry_excel_to_db(file: SpooledTemporaryFile, refresh: bool, db: Session):
    if refresh:
        refresh_table(db)

    f = await file.read()
    xlsx = io.BytesIO(f)
    workbook = load_workbook(xlsx, data_only=True)
    worksheet = workbook.worksheets[0]

    for row in worksheet.iter_rows(min_row=2):
        schema = IndustryCreationSchema(name=row[0].value)
        add_industry(db, schema)


def parse_industry(filename: str, db: Session, only_first: Union[int, None] = None):
    if only_first is not None:
        only_first += 2

    file_path = os.path.join(DATA_FOLDER_PATH, filename)

    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active

    salary_dict = dict()

    for row in worksheet.iter_rows(min_row=2, max_row=only_first, max_col=6, values_only=True):
        if db.query(IndustryModel).filter(IndustryModel.name == row[1]).first():
            continue

        schema = IndustryCreationSchema(name=row[1], avg_salary=0.0)

        if not str(row[1]) in salary_dict:
            salary_dict[row[1]] = []
        salary_dict[row[1]].append(float(row[5]))

        res = add_industry(db, schema)

    for industry_name, salaries in salary_dict.items():
        avg = sum(salaries) / len(salaries)
        industry = db.query(IndustryModel).filter(IndustryModel.name == industry_name).first()
        industry.avg_salary = avg
        db.commit()
