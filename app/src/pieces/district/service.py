import io
from tempfile import SpooledTemporaryFile

from sqlalchemy import func, text
from sqlalchemy.orm import Session

from app.src.config import DATA_FOLDER_PATH
from app.src.pieces.district.models import DistrictModel
from app.src.pieces.district.schemas import DistrictCreationSchema
import os
from typing import Union

from openpyxl import load_workbook
from sqlalchemy.orm import Session


def refresh_table(db: Session):
    db.execute(text("TRUNCATE TABLE district"))


def get_district_by_id(db: Session, additional_service_id: int) -> DistrictModel:
    return db.query(DistrictModel) \
        .filter(DistrictModel.id == additional_service_id).first()


def get_districts(db: Session, skip: int = 0, limit: int = 100) -> list[DistrictModel]:
    return db.query(DistrictModel).offset(skip).limit(limit).all()


def get_district_suggestions(db: Session, subtext: str = '',
                             skip: int = 0, limit: int = 100) -> list[DistrictModel]:
    if subtext == '':
        return get_districts(db, skip, limit)
    return db.query(DistrictModel) \
        .filter(func.lower(DistrictModel.name).contains(subtext.lower())).offset(skip).limit(limit).all()


def add_district(db: Session, district: DistrictCreationSchema) -> DistrictModel:
    district = DistrictModel(**district.dict())
    db.add(district)
    db.commit()
    db.refresh(district)
    return district


def update_district(db: Session, id: int,
                    schema: DistrictCreationSchema) -> DistrictModel:
    district = db.query(DistrictModel) \
        .filter(DistrictModel.id == id).first()
    district.average_price_per_m2_rub = schema.average_price_per_m2_rub
    district.name = schema.name
    db.commit()
    return district


def delete_district(db: Session, id: int, ) -> DistrictModel:
    district = db.query(DistrictModel) \
        .filter(DistrictModel.id == id).first()
    db.delete(district)
    db.commit()
    return district


async def upload_district_excel_to_db(file: SpooledTemporaryFile, refresh: bool, db: Session):
    if refresh:
        refresh_table(db)

    f = await file.read()
    xlsx = io.BytesIO(f)
    workbook = load_workbook(xlsx, data_only=True)
    worksheet = workbook.worksheets[0]

    for row in worksheet.iter_rows(min_row=2):
        schema = DistrictCreationSchema(name=row[0].value, average_price_per_m2_rub=float(row[1].value))
        add_district(db, schema)


def parse_district(filename: str, db: Session, only_first: Union[int, None] = None):
    if only_first is not None:
        only_first += 2

    file_path = os.path.join(DATA_FOLDER_PATH, filename)

    workbook = load_workbook(file_path, data_only=True)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, max_row=only_first, max_col=5, values_only=True):
        if row[2] is None:
            continue
        schema = DistrictCreationSchema(name=row[1], average_price_per_m2_rub=float(row[2]))
        res = add_district(db, schema)
        # print('eq created')
        # print(res.average_price_per_m2_rub)
        # print(res.name)
        # print(res.id)
